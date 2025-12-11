# app/routers/reports.py
from io import BytesIO
from datetime import date, datetime
from typing import Optional

from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse

from app.db import invoices
from app.routers.invoices import _parse_date
from app.auth import get_current_user

router = APIRouter(prefix="/reports", tags=["reports"])


def _customer_name_from_snapshot(snap: dict | None) -> str:
    snap = snap or {}
    company = (snap.get("company_name") or "").strip()
    if company:
        return company
    first = (snap.get("first_name") or "").strip()
    last = (snap.get("last_name") or "").strip()
    return f"{first} {last}".strip()


async def _collect_finance(date_from: str | None, date_to: str | None, owner_id):
    """
    Raccoglie KPI e righe fatture scoperte per owner_id e, se presenti,
    nel range data [date_from, date_to] (formato YYYY-MM-DD).
    """
    q: dict = {"owner_id": owner_id}
    df = _parse_date(date_from)
    dt = _parse_date(date_to)
    if df or dt:
        rng = {}
        if df:
            rng["$gte"] = df
        if dt:
            rng["$lte"] = dt
        q["issue_date"] = rng

    # KPI somme (per utente + range)
    pipeline_sum = [{"$match": q}]
    pipeline_sum += [
        {
            "$group": {
                "_id": None,
                "sum_total": {"$sum": {"$ifNull": ["$total", 0]}},
                "sum_paid": {
                    "$sum": {
                        "$cond": [
                            {"$eq": [{"$ifNull": ["$paid", False]}, True]},
                            {"$ifNull": ["$total", 0]},
                            0,
                        ]
                    }
                },
            }
        }
    ]
    sums = await invoices.aggregate(pipeline_sum).to_list(1)
    sum_total = float(sums[0]["sum_total"]) if sums else 0.0
    sum_paid = float(sums[0]["sum_paid"]) if sums else 0.0

    # elenco fatture (per sheet "Invoices")
    cur = invoices.find(q).sort([("issue_date", 1), ("_id", 1)])
    rows = []
    async for d in cur:
        rows.append(
            {
                "number": d.get("number"),
                "issue_date": d.get("issue_date"),
                "due_date": d.get("due_date"),
                "customer": _customer_name_from_snapshot(d.get("customer_snapshot")),
                "status": d.get("status"),
                "paid": bool(d.get("paid")),
                "paid_at": d.get("paid_at"),
                "subtotal": float(d.get("subtotal", 0)),
                "vat_total": float(d.get("vat_total", 0)),
                "total": float(d.get("total", 0)),
            }
        )

    return {
        "kpi": {
            "total": round(sum_total, 2),
            "paid": round(sum_paid, 2),
            "outstanding": round(sum_total - sum_paid, 2),
        },
        "rows": rows,
    }


@router.get("/finance")
async def finance_json(
    date_from: Optional[str] = Query(None, description="YYYY-MM-DD"),
    date_to: Optional[str] = Query(None, description="YYYY-MM-DD"),
    user=Depends(get_current_user),
):
    return await _collect_finance(date_from, date_to, user["_id"])


@router.get("/finance/export.xlsx")
async def finance_export_xlsx(
    date_from: Optional[str] = Query(None, description="YYYY-MM-DD"),
    date_to: Optional[str] = Query(None, description="YYYY-MM-DD"),
    user=Depends(get_current_user),
):
    """
    Esporta un Excel con:
    - KPI (totale, incassato, da incassare)
    - Invoices (tabellare)
    Il tutto scoperchiato per l'utente corrente.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except Exception:
        raise HTTPException(status_code=500, detail="openpyxl non installato. Esegui: pip install openpyxl")

    data = await _collect_finance(date_from, date_to, user["_id"])

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "KPI"

    ws1["A1"] = "Report finanziario"
    ws1["A1"].font = Font(bold=True, size=14)
    ws1.merge_cells("A1:C1")

    ws1["A3"], ws1["B3"] = "Periodo FROM", date_from or "-"
    ws1["A4"], ws1["B4"] = "Periodo TO", date_to or "-"
    ws1["A6"], ws1["B6"] = "Totale emesso", data["kpi"]["total"]
    ws1["A7"], ws1["B7"] = "Incassato", data["kpi"]["paid"]
    ws1["A8"], ws1["B8"] = "Da incassare", data["kpi"]["outstanding"]

    ws2 = wb.create_sheet("Invoices")
    headers = [
        "Numero",
        "Data",
        "Scadenza",
        "Cliente",
        "Stato",
        "Pagata",
        "Data Pag.",
        "Imponibile",
        "IVA",
        "Totale",
    ]
    ws2.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws2.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="EEEEEE")

    for r in data["rows"]:
        ws2.append(
            [
                r["number"],
                r["issue_date"],
                r["due_date"],
                r["customer"],
                r["status"],
                "Sì" if r["paid"] else "No",
                r["paid_at"],
                r["subtotal"],
                r["vat_total"],
                r["total"],
            ]
        )

    # auto-fit semplice
    for col in ws2.columns:
        max_len = 10
        for cell in col:
            val = cell.value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"report_finance_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
